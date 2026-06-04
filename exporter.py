"""Turn a scrape() result into a multi-sheet Excel workbook (in memory)."""
from __future__ import annotations

import io
from datetime import datetime, timezone, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ClickUp timestamps are UTC ms; show them in Bangkok time.
_TZ = timezone(timedelta(hours=7))

_HEADER_FILL = PatternFill("solid", fgColor="2D3A8C")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


# --------------------------------------------------------------------------
# value helpers
# --------------------------------------------------------------------------
def ms_to_str(value) -> str:
    """Convert a ClickUp ms timestamp (str/int) to 'YYYY-MM-DD HH:MM'."""
    if value in (None, "", 0, "0"):
        return ""
    try:
        ms = int(value)
    except (TypeError, ValueError):
        return str(value)
    try:
        return datetime.fromtimestamp(ms / 1000, _TZ).strftime("%Y-%m-%d %H:%M")
    except (OverflowError, OSError, ValueError):
        return str(value)


def content_to_text(content) -> str:
    """Flatten a ClickUp rich-text value into plain text.

    `content`/`comment` are Quill-delta-like: either a dict {"ops":[...]} or a
    list of segments [{"text": "...", "insert": ...}, ...].
    """
    if content is None:
        return ""
    if isinstance(content, str):
        return content
    ops = content.get("ops", []) if isinstance(content, dict) else content
    parts = []
    for seg in ops or []:
        if not isinstance(seg, dict):
            parts.append(str(seg))
            continue
        ins = seg.get("insert", seg.get("text"))
        if isinstance(ins, str):
            parts.append(ins)
        elif isinstance(ins, dict):
            # embeds: mentions, links, images
            label = (ins.get("link_mention", {}).get("url")
                     or ins.get("link", {}).get("url")
                     or ins.get("user", {}).get("username")
                     or ins.get("tag", {}).get("text"))
            if label:
                parts.append(f"[{label}]")
    return "".join(parts).strip()


def task_description(task: dict) -> str:
    return (task.get("text_content")
            or content_to_text(task.get("content"))
            or "")


def names(items, key="username") -> str:
    out = []
    for it in items or []:
        if isinstance(it, dict):
            out.append(it.get(key) or it.get("name") or it.get("email") or "")
        else:
            out.append(str(it))
    return ", ".join(x for x in out if x)


def priority_label(task: dict) -> str:
    p = task.get("priority")
    if isinstance(p, dict):
        return p.get("priority", "")
    return p or ""


def status_label(task: dict) -> str:
    s = task.get("status")
    if isinstance(s, dict):
        return s.get("status", "")
    return s or ""


def task_url(task: dict) -> str:
    tid = task.get("id")
    return f"https://app.clickup.com/t/{tid}" if tid else ""


def custom_field_value(cf: dict):
    """Best-effort readable value for a custom field entry."""
    val = cf.get("value")
    cf_type = cf.get("type")
    cfg = cf.get("type_config", {}) or {}
    if val is None:
        return ""
    if cf_type == "drop_down":
        options = cfg.get("options", [])
        try:
            idx = int(val)
            for o in options:
                if o.get("orderindex") == idx or o.get("id") == val:
                    return o.get("name", val)
        except (TypeError, ValueError):
            pass
        for o in options:
            if o.get("id") == val:
                return o.get("name", val)
        return val
    if cf_type == "labels" and isinstance(val, list):
        options = {o.get("id"): o.get("label", o.get("name")) for o in cfg.get("options", [])}
        return ", ".join(str(options.get(v, v)) for v in val)
    if cf_type in ("users",) and isinstance(val, list):
        return names(val)
    if cf_type == "date":
        return ms_to_str(val)
    if isinstance(val, (list, dict)):
        return str(val)
    return val


# --------------------------------------------------------------------------
# sheet writers
# --------------------------------------------------------------------------
def _style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def _autosize(ws, max_width=60):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        longest = 0
        for cell in col:
            v = cell.value
            if v is None:
                continue
            longest = max(longest, max((len(line) for line in str(v).split("\n")), default=0))
        ws.column_dimensions[letter].width = min(max(12, longest + 2), max_width)


def _write_rows(ws, headers, rows):
    ws.append(headers)
    for row in rows:
        ws.append(row)
    _style_header(ws, len(headers))
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).alignment = Alignment(
                vertical="top", wrap_text=True)
    _autosize(ws)


def _collect_custom_field_names(tasks):
    seen, order = set(), []
    for t in tasks:
        for cf in t.get("custom_fields", []) or t.get("customFields", []) or []:
            name = cf.get("name")
            if name and name not in seen:
                seen.add(name)
                order.append(name)
    return order


def build_workbook(result: dict) -> bytes:
    tasks = result.get("tasks", [])
    wb = Workbook()

    cf_names = _collect_custom_field_names(tasks)

    # ---- Tasks sheet -----------------------------------------------------
    ws = wb.active
    ws.title = "Tasks"
    base_headers = [
        "Task ID", "Custom ID", "Task Name", "Status", "Status Type",
        "Priority", "Assignees", "Tags", "Due Date", "Start Date",
        "Created", "Updated", "Closed", "Done",
        "Creator", "Time Estimate", "Points",
        "Comments", "Attachments", "Subtasks", "Checklists",
        "Description", "Parent", "List / Project", "URL",
    ]
    headers = base_headers + [f"CF: {n}" for n in cf_names]

    rows = []
    for t in tasks:
        cf_map = {}
        for cf in t.get("custom_fields", []) or t.get("customFields", []) or []:
            cf_map[cf.get("name")] = custom_field_value(cf)
        proj = t.get("project") or {}
        row = [
            t.get("id", ""),
            t.get("custom_id") or "",
            t.get("name", ""),
            status_label(t),
            (t.get("status") or {}).get("type", "") if isinstance(t.get("status"), dict) else "",
            priority_label(t),
            names(t.get("assignees")),
            names(t.get("tags"), key="name"),
            ms_to_str(t.get("due_date")),
            ms_to_str(t.get("start_date")),
            ms_to_str(t.get("date_created")),
            ms_to_str(t.get("date_updated")),
            ms_to_str(t.get("date_closed")),
            ms_to_str(t.get("date_done")),
            (t.get("creator") or {}).get("username", ""),
            t.get("time_estimate_string") or "",
            t.get("points") if t.get("points") is not None else "",
            len(t.get("comments", []) or []),
            t.get("attachments_count", len(t.get("attachments", []) or [])),
            t.get("subtasks_count", len(t.get("subtasks", []) or [])),
            len(t.get("checklists", []) or []),
            task_description(t),
            (t.get("parent_task") or {}).get("name", "") if isinstance(t.get("parent_task"), dict) else (t.get("parent") or ""),
            proj.get("name", "") if isinstance(proj, dict) else "",
            task_url(t),
        ]
        row += [cf_map.get(n, "") for n in cf_names]
        rows.append(row)
    _write_rows(ws, headers, rows)

    # ---- Comments sheet --------------------------------------------------
    ws_c = wb.create_sheet("Comments")
    crows = []
    for t in tasks:
        tname = t.get("name", "")
        tid = t.get("id", "")
        for c in t.get("comments", []) or []:
            user = (c.get("user") or {}).get("username", "")
            text = c.get("text_content") or content_to_text(c.get("comment"))
            reactions = ", ".join(
                f"{r.get('reaction')}({len(r.get('user_ids', []) or r.get('users', []) or [])})"
                for r in c.get("reactions", []) or [])
            crows.append([
                tid, tname,
                c.get("comment_number", ""),
                user,
                ms_to_str(c.get("date")),
                ms_to_str(c.get("date_updated")),
                text,
                c.get("threaded_replies", c.get("new_thread_count", 0)),
                reactions,
            ])
    _write_rows(ws_c, [
        "Task ID", "Task Name", "Comment #", "Author", "Date", "Edited",
        "Comment", "Replies", "Reactions",
    ], crows)

    # ---- Subtasks sheet --------------------------------------------------
    ws_s = wb.create_sheet("Subtasks")
    srows = []
    for t in tasks:
        for st in t.get("subtasks", []) or []:
            srows.append([
                t.get("id", ""), t.get("name", ""),
                st.get("id", ""), st.get("name", ""),
                status_label(st),
                names(st.get("assignees")),
                ms_to_str(st.get("due_date")),
                st.get("time_estimate_string") or "",
            ])
    _write_rows(ws_s, [
        "Parent Task ID", "Parent Task", "Subtask ID", "Subtask Name",
        "Status", "Assignees", "Due Date", "Time Estimate",
    ], srows)

    # ---- Checklist items sheet ------------------------------------------
    ws_k = wb.create_sheet("Checklists")
    krows = []
    for t in tasks:
        for cl in t.get("checklists", []) or []:
            cl_name = cl.get("name", "")
            items = cl.get("items", []) or []
            if not items:
                krows.append([t.get("id", ""), t.get("name", ""), cl_name, "", ""])
            for it in items:
                krows.append([
                    t.get("id", ""), t.get("name", ""), cl_name,
                    it.get("name", ""),
                    "✓" if it.get("resolved") else "",
                ])
    _write_rows(ws_k, [
        "Task ID", "Task Name", "Checklist", "Item", "Done",
    ], krows)

    # ---- Attachments sheet ----------------------------------------------
    ws_a = wb.create_sheet("Attachments")
    arows = []
    for t in tasks:
        for at in t.get("attachments", []) or []:
            size = at.get("size")
            size_kb = f"{round(size / 1024, 1)} KB" if isinstance(size, (int, float)) else ""
            arows.append([
                t.get("id", ""), t.get("name", ""),
                at.get("title", ""),
                at.get("extension", "") or at.get("mimetype", ""),
                size_kb,
                ms_to_str(at.get("date")),
                at.get("url", ""),
            ])
    _write_rows(ws_a, [
        "Task ID", "Task Name", "File", "Type", "Size", "Uploaded", "URL",
    ], arows)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
